import csv
import unicodedata
from io import BytesIO
from datetime import date

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.utils.dateparse import parse_date, parse_time
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


from .forms import DriverForm, DriverGroupForm, LoginPasswordForm, ShiftFilterForm, ShiftRecordForm, VehicleForm, TIME_CHOICES
from .models import Driver, DriverGroup, ShiftRecord, Vehicle
from .permissions import coordinator_or_manager_required, is_manager, manager_required


def role_login(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    selected_role = request.POST.get('role') or request.GET.get('role') or ''
    form = LoginPasswordForm(initial={'role': selected_role})

    if request.method == 'POST':
        form = LoginPasswordForm(request.POST)
        if form.is_valid():
            role = form.cleaned_data['role']
            password = form.cleaned_data['password']
            username = 'mudur' if role == 'mudur' else 'koordine'
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
            messages.error(request, 'Şifre hatalı. Lütfen tekrar deneyin.')

    return render(request, 'panel/login.html', {'form': form, 'selected_role': selected_role})


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
@coordinator_or_manager_required
def dashboard(request):
    today = date.today()
    context = {
        'active_drivers': Driver.objects.filter(is_active=True).count(),
        'active_vehicles': Vehicle.objects.filter(is_active=True).count(),
        'today_records': ShiftRecord.objects.filter(operation_date=today).count(),
        'last_records': ShiftRecord.objects.select_related('driver', 'vehicle').order_by('-created_at')[:8],
    }
    return render(request, 'panel/dashboard.html', context)


@login_required
@coordinator_or_manager_required
def driver_management(request):
    add_form = DriverForm()
    group_form = DriverGroupForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_driver':
            add_form = DriverForm(request.POST)
            if add_form.is_valid():
                add_form.save()
                messages.success(request, 'Sürücü eklendi.')
                return redirect('drivers')
        elif action == 'add_group':
            group_form = DriverGroupForm(request.POST)
            if group_form.is_valid():
                group_form.save()
                messages.success(request, 'Grup eklendi.')
                return redirect('drivers')

    query = request.GET.get('q', '').strip()
    group_id = request.GET.get('group', '').strip()
    status = request.GET.get('status', 'active')
    drivers = Driver.objects.select_related('group')
    if query:
        drivers = drivers.filter(full_name__icontains=query)
    if group_id:
        drivers = drivers.filter(group_id=group_id)
    if status == 'active':
        drivers = drivers.filter(is_active=True)
    elif status == 'passive':
        drivers = drivers.filter(is_active=False)

    context = {
        'drivers': drivers.order_by('full_name'),
        'groups': DriverGroup.objects.all(),
        'add_form': add_form,
        'group_form': group_form,
        'query': query,
        'selected_group': group_id,
        'status': status,
    }
    return render(request, 'panel/drivers.html', context)


@login_required
@coordinator_or_manager_required
def driver_edit(request, pk):
    driver = get_object_or_404(Driver, pk=pk)
    form = DriverForm(instance=driver)
    if request.method == 'POST':
        form = DriverForm(request.POST, instance=driver)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sürücü bilgisi güncellendi.')
            return redirect('drivers')
    return render(request, 'panel/form_page.html', {'title': 'Sürücü Düzenle', 'form': form, 'back_url': reverse('drivers')})


@login_required
@coordinator_or_manager_required
def driver_toggle(request, pk):
    driver = get_object_or_404(Driver, pk=pk)
    if request.method == 'POST':
        driver.is_active = not driver.is_active
        driver.save(update_fields=['is_active', 'updated_at'])
        messages.success(request, 'Sürücü durumu güncellendi.')
    return redirect('drivers')


@login_required
@coordinator_or_manager_required
def plate_management(request):
    form = VehicleForm()
    if request.method == 'POST':
        form = VehicleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Plaka eklendi.')
            return redirect('plates')

    query = request.GET.get('q', '').strip()
    status = request.GET.get('status', 'active')
    vehicles = Vehicle.objects.all()
    if query:
        vehicles = vehicles.filter(plate__icontains=query)
    if status == 'active':
        vehicles = vehicles.filter(is_active=True)
    elif status == 'passive':
        vehicles = vehicles.filter(is_active=False)

    return render(request, 'panel/plates.html', {
        'vehicles': vehicles.order_by('plate'),
        'form': form,
        'query': query,
        'status': status,
    })


@login_required
@coordinator_or_manager_required
def plate_edit(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    form = VehicleForm(instance=vehicle)
    if request.method == 'POST':
        form = VehicleForm(request.POST, instance=vehicle)
        if form.is_valid():
            form.save()
            messages.success(request, 'Plaka güncellendi.')
            return redirect('plates')
    return render(request, 'panel/form_page.html', {'title': 'Plaka Düzenle', 'form': form, 'back_url': reverse('plates')})


@login_required
@coordinator_or_manager_required
def plate_toggle(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    if request.method == 'POST':
        vehicle.is_active = not vehicle.is_active
        vehicle.save(update_fields=['is_active', 'updated_at'])
        messages.success(request, 'Plaka durumu güncellendi.')
    return redirect('plates')


@login_required
@coordinator_or_manager_required
def shift_entry(request):
    form = ShiftRecordForm()
    time_options = [value for value, _label in TIME_CHOICES if value]
    active_drivers = Driver.objects.filter(is_active=True).select_related('group').order_by('full_name')
    active_vehicles = Vehicle.objects.filter(is_active=True).order_by('plate')

    if request.method == 'POST':
        action = request.POST.get('action', 'single_shift')

        if action == 'single_shift':
            form = ShiftRecordForm(request.POST)
            if form.is_valid():
                record = form.save(commit=False)
                record.created_by = request.user
                record.updated_by = request.user
                record.save()
                messages.success(request, 'Vardiya kaydı oluşturuldu.')
                return redirect('shift_entry')
            else:
                messages.error(request, 'Kayıt oluşturulmadı. Tarih boş bırakılırsa bugün seçilir; ancak sürücü, vardiya saatleri ve plaka zorunludur.')

        elif action == 'bulk_shift':
            bulk_date = parse_date(request.POST.get('bulk_operation_date', '').strip()) or date.today()
            try:
                row_count = int(request.POST.get('bulk_row_count', '10'))
            except ValueError:
                row_count = 10
            row_count = max(1, min(row_count, 200))

            created_count = 0
            skipped_count = 0
            error_rows = []

            for index in range(row_count):
                driver_id = request.POST.get(f'bulk_driver_{index}', '').strip()
                shift_start = parse_time(request.POST.get(f'bulk_shift_start_{index}', '').strip())
                shift_end = parse_time(request.POST.get(f'bulk_shift_end_{index}', '').strip())
                pickup_time = parse_time(request.POST.get(f'bulk_pickup_{index}', '').strip())
                dropoff_time = parse_time(request.POST.get(f'bulk_dropoff_{index}', '').strip())
                vehicle_id = request.POST.get(f'bulk_vehicle_{index}', '').strip()
                note = request.POST.get(f'bulk_note_{index}', '').strip()

                row_has_data = any([driver_id, vehicle_id, note])
                if not row_has_data:
                    continue

                if not bulk_date or not driver_id or not shift_start or not shift_end:
                    skipped_count += 1
                    error_rows.append(str(index + 1))
                    continue

                try:
                    driver = Driver.objects.get(pk=driver_id, is_active=True)
                except Driver.DoesNotExist:
                    skipped_count += 1
                    error_rows.append(str(index + 1))
                    continue

                vehicle = None
                if not vehicle_id:
                    skipped_count += 1
                    error_rows.append(str(index + 1))
                    continue

                try:
                    vehicle = Vehicle.objects.get(pk=vehicle_id, is_active=True)
                except Vehicle.DoesNotExist:
                    skipped_count += 1
                    error_rows.append(str(index + 1))
                    continue

                ShiftRecord.objects.create(
                    operation_date=bulk_date,
                    driver=driver,
                    shift_start=shift_start,
                    shift_end=shift_end,
                    vehicle_pickup_time=pickup_time,
                    vehicle_dropoff_time=dropoff_time,
                    vehicle=vehicle,
                    note=note,
                    created_by=request.user,
                    updated_by=request.user,
                )
                created_count += 1

            if created_count:
                messages.success(request, f'{created_count} toplu vardiya kaydı oluşturuldu.')
            if skipped_count:
                rows = ', '.join(error_rows[:12])
                extra = '...' if len(error_rows) > 12 else ''
                messages.warning(request, f'{skipped_count} satır eksik/hatalı olduğu için kaydedilmedi. Sürücü, giriş-çıkış saati ve plaka zorunludur. Satırlar: {rows}{extra}')
            if not created_count and not skipped_count:
                messages.warning(request, 'Toplu girişte kaydedilecek satır bulunamadı.')
            return redirect('shift_entry')

    records = ShiftRecord.objects.select_related('driver', 'vehicle').order_by('-created_at')[:20]
    return render(request, 'panel/shift_entry.html', {
        'form': form,
        'records': records,
        'time_options': time_options,
        'active_drivers': active_drivers,
        'active_vehicles': active_vehicles,
        'bulk_rows': range(8),
        'today_iso': date.today().isoformat(),
    })


def _filtered_records(request):
    records = ShiftRecord.objects.select_related('driver', 'driver__group', 'vehicle').all()
    form = ShiftFilterForm(request.GET or None)
    if form.is_valid():
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        driver = form.cleaned_data.get('driver')
        vehicle = form.cleaned_data.get('vehicle')
        search = form.cleaned_data.get('search')
        if start_date:
            records = records.filter(operation_date__gte=start_date)
        if end_date:
            records = records.filter(operation_date__lte=end_date)
        if driver:
            records = records.filter(driver=driver)
        if vehicle:
            records = records.filter(vehicle=vehicle)
        if search:
            records = records.filter(Q(driver__full_name__icontains=search) | Q(vehicle__plate__icontains=search) | Q(note__icontains=search))
    return form, records


@login_required
@manager_required
def shift_logs(request):
    form, records = _filtered_records(request)
    return render(request, 'panel/shift_logs.html', {'form': form, 'records': records[:500]})


@login_required
@manager_required
def shift_edit(request, pk):
    record = get_object_or_404(ShiftRecord, pk=pk)
    form = ShiftRecordForm(instance=record)
    if request.method == 'POST':
        form = ShiftRecordForm(request.POST, instance=record)
        if form.is_valid():
            item = form.save(commit=False)
            item.updated_by = request.user
            item.save()
            messages.success(request, 'Vardiya kaydı güncellendi.')
            return redirect('shift_logs')
    return render(request, 'panel/form_page.html', {'title': 'Vardiya Kaydı Düzenle', 'form': form, 'back_url': reverse('shift_logs')})


@login_required
@manager_required
def shift_delete(request, pk):
    record = get_object_or_404(ShiftRecord, pk=pk)
    if request.method == 'POST':
        record.delete()
        messages.success(request, 'Vardiya kaydı silindi.')
        return redirect('shift_logs')
    return render(request, 'panel/confirm_delete.html', {'object': record, 'back_url': reverse('shift_logs')})


def _add_bar_percent(items):
    max_total = max([item.get('total', 0) for item in items], default=0) or 1
    for item in items:
        item['percent'] = round((item.get('total', 0) / max_total) * 100, 2)
    return items


@login_required
@manager_required
def analysis_reports(request):
    form, records = _filtered_records(request)
    qs = records.order_by('-operation_date', 'shift_start')[:1000]
    driver_summary = records.values('driver__full_name').annotate(total=Count('id')).order_by('-total')[:20]
    vehicle_summary = records.values('vehicle__plate').annotate(total=Count('id')).order_by('-total')[:20]
    group_summary = records.values('driver__group__name').annotate(total=Count('id')).order_by('-total')[:20]

    daily_chart = [
        {'label': row['operation_date'].strftime('%d.%m.%Y'), 'total': row['total']}
        for row in records.values('operation_date').annotate(total=Count('id')).order_by('operation_date')[:60]
    ]
    shift_chart = [
        {
            'label': f"{row['shift_start'].strftime('%H:%M')} - {row['shift_end'].strftime('%H:%M')}",
            'total': row['total'],
        }
        for row in records.values('shift_start', 'shift_end').annotate(total=Count('id')).order_by('-total')[:20]
    ]
    vehicle_chart = [
        {'label': row['vehicle__plate'] or 'Plakasız', 'total': row['total']}
        for row in records.values('vehicle__plate').annotate(total=Count('id')).order_by('-total')[:20]
    ]
    group_chart = [
        {'label': row['driver__group__name'] or 'Grupsuz', 'total': row['total']}
        for row in records.values('driver__group__name').annotate(total=Count('id')).order_by('-total')[:20]
    ]

    return render(request, 'panel/analysis.html', {
        'form': form,
        'records': qs,
        'driver_summary': driver_summary,
        'vehicle_summary': vehicle_summary,
        'group_summary': group_summary,
        'daily_chart': _add_bar_percent(daily_chart),
        'shift_chart': _add_bar_percent(shift_chart),
        'vehicle_chart': _add_bar_percent(vehicle_chart),
        'group_chart': _add_bar_percent(group_chart),
    })



def _export_filename(extension):
    return f"istclbdriver_vardiya_kayitlari_{date.today().strftime('%Y%m%d')}.{extension}"


def _plain_text(value):
    """PDF icin Turkce karakterleri guvenli ASCII karsiliklarina cevirir."""
    if value is None:
        return ''
    value = str(value)
    translations = str.maketrans({
        'Ç': 'C', 'ç': 'c', 'Ğ': 'G', 'ğ': 'g', 'İ': 'I', 'ı': 'i',
        'Ö': 'O', 'ö': 'o', 'Ş': 'S', 'ş': 's', 'Ü': 'U', 'ü': 'u',
    })
    value = value.translate(translations)
    return unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')


def _record_export_rows(records):
    rows = []
    for r in records.select_related('driver', 'driver__group', 'vehicle').order_by('operation_date', 'shift_start', 'driver__full_name'):
        rows.append({
            'Tarih': r.operation_date.strftime('%d.%m.%Y'),
            'Sürücü': r.driver.full_name,
            'Grup': r.driver.group.name if r.driver.group else '',
            'Vardiya Giriş': r.shift_start.strftime('%H:%M') if r.shift_start else '',
            'Vardiya Çıkış': r.shift_end.strftime('%H:%M') if r.shift_end else '',
            'Araç Alma': r.vehicle_pickup_time.strftime('%H:%M') if r.vehicle_pickup_time else '',
            'Araç Bırakma': r.vehicle_dropoff_time.strftime('%H:%M') if r.vehicle_dropoff_time else '',
            'Plaka': r.vehicle.plate if r.vehicle else 'PLAKA GIRILMEDI',
            'Not': r.note or '',
        })
    return rows


@login_required
@manager_required
def export_shift_csv(request):
    _form, records = _filtered_records(request)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{_export_filename("csv")}"'
    response.write('\ufeff')
    writer = csv.writer(response)
    headers = ['Tarih', 'Sürücü', 'Grup', 'Vardiya Giriş', 'Vardiya Çıkış', 'Araç Alma', 'Araç Bırakma', 'Plaka', 'Not']
    writer.writerow(headers)
    for row in _record_export_rows(records):
        writer.writerow([row[h] for h in headers])
    return response


@login_required
@manager_required
def export_shift_excel(request):
    _form, records = _filtered_records(request)
    headers = ['Tarih', 'Sürücü', 'Grup', 'Vardiya Giriş', 'Vardiya Çıkış', 'Araç Alma', 'Araç Bırakma', 'Plaka', 'Not']
    rows = _record_export_rows(records)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vardiya Kayıtları'

    ws.merge_cells('A1:I1')
    ws['A1'] = 'ISTCLB Driver - Vardiya Kayıtları'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='0B1F3A')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append([])
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    thin = Side(style='thin', color='D8DEE9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[3]:
        cell.font = Font(bold=True, color='0B1F3A')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in rows:
        ws.append([row[h] for h in headers])

    for row_cells in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
        if row_cells[7].value == 'PLAKA GIRILMEDI':
            row_cells[7].font = Font(color='9C0006', bold=True)
            row_cells[7].fill = PatternFill('solid', fgColor='FFC7CE')

    widths = [14, 28, 22, 15, 15, 14, 16, 18, 36]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:I{ws.max_row}'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_export_filename("xlsx")}"'
    return response


@login_required
@manager_required
def export_shift_pdf(request):
    _form, records = _filtered_records(request)
    headers = ['Tarih', 'Surucu', 'Grup', 'Giris', 'Cikis', 'Arac Alma', 'Arac Birakma', 'Plaka', 'Not']
    rows = _record_export_rows(records)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph('ISTCLB Driver - Vardiya Kayitlari', styles['Title']),
        Spacer(1, 8),
    ]

    table_data = [headers]
    for row in rows[:1200]:
        table_data.append([
            _plain_text(row['Tarih']),
            _plain_text(row['Sürücü']),
            _plain_text(row['Grup']),
            _plain_text(row['Vardiya Giriş']),
            _plain_text(row['Vardiya Çıkış']),
            _plain_text(row['Araç Alma']),
            _plain_text(row['Araç Bırakma']),
            _plain_text(row['Plaka']),
            _plain_text(row['Not'])[:80],
        ])

    table = Table(table_data, repeatRows=1, colWidths=[52, 120, 86, 48, 48, 58, 68, 72, 160])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B1F3A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D8DEE9')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FBFF')]),
    ]))
    elements.append(table)
    if len(rows) > 1200:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph('Not: PDF cikti performans icin ilk 1200 kaydi icerir. Tum veri icin Excel ciktisini kullanin.', styles['Normal']))
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_export_filename("pdf")}"'
    return response
